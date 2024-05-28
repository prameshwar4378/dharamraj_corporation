from django.shortcuts import render,redirect,get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .forms import *
from Developer.models import *
from .filters import *
from django.contrib import messages
from django.http import Http404
# Create your views here.
from django.db.models import Sum, F, ExpressionWrapper, fields, Q
from datetime import date,timedelta,datetime, timedelta
from django.db.models.functions import Coalesce
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.decorators import login_required, user_passes_test 

from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.core.cache import cache

def is_admin(user):
    return user.is_authenticated and hasattr(user, 'is_admin') and user.is_admin

def admin_required(view_func):
    decorated_view_func = login_required(user_passes_test(is_admin, login_url='/accounts/login')(view_func))
    return decorated_view_func


@admin_required
def dashboard(request):
    try:

        today_date = date.today()
        end_of_day = today_date + timedelta(days=1)
        total_grand_total_today = Invoice.objects.filter(invoice_date=today_date).aggregate(Sum('grand_total'))['grand_total__sum'] or 0
        total_gst_amount_today = Invoice.objects.filter(invoice_date=today_date).aggregate(Sum('total_gst_amount'))['total_gst_amount__sum'] or 0
        total_quantity_today = InvoiceItem.objects.filter(invoice__invoice_date=today_date).aggregate(Sum('quantity'))['quantity__sum'] or 0
        total_invoices_today = Invoice.objects.filter(invoice_date=today_date).count()

        total_credited_amount_today = Account.objects.filter(
            is_credit=True, 
            transaction_date__range=(today_date, end_of_day)
        ).aggregate(Sum('amount'))['amount__sum'] or 0

        total_debited_amount_today = Account.objects.filter(
            is_credit=False, 
            transaction_date__range=(today_date, end_of_day)
        ).aggregate(Sum('amount'))['amount__sum'] or 0

        total_credited_amount = Account.objects.filter(is_credit=True).aggregate(Sum('amount'))['amount__sum'] or 0
        total_debited_amount = Account.objects.filter(is_credit=False).aggregate(Sum('amount'))['amount__sum'] or 0
        total_balance = Account.objects.aggregate(
            total_balance=Sum(ExpressionWrapper(
                F('amount') * F('is_credit') - F('amount') * ~F('is_credit'),
                output_field=fields.FloatField()
            ))
        )['total_balance'] or 0

        count_product_out_of_stock = Product.objects.filter(available_stock=0).count()
        total_available_stock = Product.objects.aggregate(total_available_stock=Sum('available_stock'))['total_available_stock'] or 0
        count_min_stock = Product.objects.filter(~Q(available_stock=0),available_stock__lt=F('minimum_stock')).count()
        count_available_stock = Product.objects.filter(available_stock__gte=F('minimum_stock')).count()
         
        start_date = date.today() - timedelta(days=30)
        top_five_dealers = Invoice.objects.filter(invoice_date__gte=start_date).values('dealer__business_name').annotate(total_amount=Coalesce(Sum('grand_total'), 0)).order_by('-total_amount')[:5]
        top_five_dealer_total_amount = [item['total_amount'] for item in top_five_dealers]
        top_five_dealer_name = [item['dealer__business_name'] for item in top_five_dealers]

        top_five_sale_products = InvoiceItem.objects.filter(invoice__invoice_date__gte=start_date).values('product__product_name').annotate(total_quantity=Coalesce(Sum('quantity'), 0)).order_by('-total_quantity')[:5]
        top_five_sale_product_quantity = [item['total_quantity'] for item in top_five_sale_products]
        top_five_sale_product_labels = [item['product__product_name'] for item in top_five_sale_products]

        context = {
            'count_product_out_of_stock':count_product_out_of_stock,
            'count_min_stock':count_min_stock,
            'count_available_stock':count_available_stock, 
            'total_available_stock':total_available_stock,
            'total_grand_total_today': total_grand_total_today,
            'total_gst_amount_today': total_gst_amount_today,
            'total_quantity_today': total_quantity_today,
            'total_invoices_today': total_invoices_today,
            'total_credited_amount_today': total_credited_amount_today,
            'total_debited_amount_today': total_debited_amount_today,
            'top_five_dealer_total_amount': top_five_dealer_total_amount,
            'top_five_dealer_name': top_five_dealer_name,
            'top_five_sale_product_quantity': top_five_sale_product_quantity,
            'top_five_sale_product_labels': top_five_sale_product_labels,
            'total_credited_amount': total_credited_amount,
            'total_debited_amount': total_debited_amount,
            'total_balance': total_balance,
        }
        return render(request, 'admin__dashboard.html', context)

    except ObjectDoesNotExist as e:
        # Handle the case where an object does not exist
        return render(request, '404.html', {'error_message': str(e)}, status=404)

    except Exception as e:
        # Handle other exceptions
        return render(request, '404.html', {'error_message': str(e)}, status=500)


@admin_required
def user_list(request):
    try:
        user_rec = CustomUser.objects.filter(is_admin=False, is_superuser=False)
        if request.method == 'POST':
            form = User_Creation(request.POST, request.FILES)
            if form.is_valid():
                fm = form.save(commit=False)
                fm.is_staff = True  # Associate the form with the specific invoice
                fm.save()
                messages.success(request, 'User Created Successfully ...!')
                return redirect('/admin/user_list/')
            else:
                # Form is not valid, handle the error
                messages.error(request, 'Error in form submission. Please check the form.')
        else:
            form = User_Creation()
        return render(request, 'admin__user_list.html', {'form': form, 'user_rec': user_rec})

    except Exception as e:
        # Handle other exceptions
        messages.error(request, f'An error occurred: {str(e)}')
        return render(request, '404.html', {'error_message': str(e)}, status=500)
    
  
@admin_required
def update_user(request, id):
    try:
        pi = get_object_or_404(CustomUser, pk=id)
        if request.method == "POST":
            fm = User_Updation(request.POST, request.FILES, instance=pi)
            if fm.is_valid():
                fm.save()
                messages.success(request, 'User Updated Successfully')
                return redirect('/admin/user_list/')
            else:
                # Form is not valid, handle the error
                messages.error(request, 'Error in form submission. Please check the form.')
        else:
            fm = User_Updation(instance=pi)

        return render(request, 'admin__update_user.html', {'form': fm})
 
    except Exception as e:
        # Handle other exceptions
        messages.error(request, f'An error occurred: {str(e)}')
        return render(request, '404.html', status=404)
    
  
@admin_required
def delete_user(request, id):
    try:
        user = get_object_or_404(CustomUser, id=id)
        user.delete()
        messages.success(request, 'User Deleted Successfully.')
        return redirect('/admin/user_list/')
    except Http404:
        return render(request, '404.html')


@admin_required
def dealer_list(request):
    try:
        dealer_rec = Dealer.objects.select_related().order_by('dealer_name')

        # Add pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(dealer_rec, 25)  # Show 10 dealers per page
        try:
            dealer_rec = paginator.page(page)
        except PageNotAnInteger:
            dealer_rec = paginator.page(1)
        except EmptyPage:
            dealer_rec = paginator.page(paginator.num_pages)

        create_dealer_fm = DealerCreationForm()
        update_dealer_fm = DealerUpdateForm()
        context = {
            'dealer_rec': dealer_rec,
            'create_dealer_fm': create_dealer_fm,
            'update_dealer_fm': update_dealer_fm
        }
        return render(request, 'admin__dealer_list.html', context)
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@admin_required
def create_dealer(request):
    try:
        form = DealerCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Dealer Created Successfully.....')
        else:
            error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
            messages.warning(request, error_message)
        return redirect('/admin/dealer_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@admin_required
def update_dealer(request): 
    try:
        if request.method == 'POST':
            dealer_id = request.POST.get('txt_id') 
            dealer_instance = get_object_or_404(Dealer, id=dealer_id)
            form = DealerUpdateForm(request.POST, instance=dealer_instance)
            if form.is_valid():
                form.save()
                messages.success(request, 'Dealer Updated Successfully.')
                return redirect('/admin/dealer_list/')
            else:
                error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
                messages.warning(request, error_message)
            return redirect('/admin/dealer_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@admin_required
def delete_dealer(request, id):
    try:
        dealer = get_object_or_404(Dealer, id=id)
        dealer.delete()
        messages.success(request, 'Dealer Deleted Successfully.')
        return redirect('/admin/dealer_list/')
    except Http404:
        return render(request, '404.html')


@admin_required
def product_list(request):
    try:
        product_rec = Product.objects.select_related()
        create_product_fm = ProductCreationForm()
        update_product_fm = ProductUpdateForm()
        add_stock = PurchaseCreationForm()
        update_stock = PurchaseUpdateForm()
        context = {'product_rec': product_rec,'add_stock':add_stock,'update_stock':update_stock, 'create_product_fm': create_product_fm, 'update_product_fm': update_product_fm}
        return render(request, 'admin__product_list.html', context)
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@admin_required
def create_product(request):
    try:
        form = ProductCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product Created Successfully.....')
        else:
            error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
            messages.warning(request, error_message)
        return redirect('/admin/product_list/')
    except Exception as e:
        # Log the exception if needed
        messages.error(request, f'An error occurred: {str(e)}')
        return render(request, '404.html')
    
@admin_required
def update_product(request): 
    try:
        if request.method == 'POST':
            product_id = request.POST.get('txt_id') 
            product = get_object_or_404(Product, id=product_id)
            form = ProductUpdateForm(request.POST, instance=product)
            if form.is_valid():
                form.save()
                messages.success(request, 'Product Updated Successfully.')
                return redirect('/admin/product_list/')
            else:
                error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
                messages.warning(request, error_message)
                return redirect('/admin/product_list/')
        else:
            return redirect('/admin/product_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@admin_required
def delete_product(request, id):
    try:
        product = get_object_or_404(Product, id=id)
        product.delete()
        messages.success(request, 'Product Deleted Successfully.')
        return redirect('/admin/product_list/')
    except Http404:
        return render(request, '404.html')

@admin_required
def purchase_list(request):
    try:
        stock_rec = Purchase.objects.select_related().order_by('-id')

        # Add pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(stock_rec, 25)  # Show 10 purchases per page
        try:
            stock_rec = paginator.page(page)
        except PageNotAnInteger:
            stock_rec = paginator.page(1)
        except EmptyPage:
            stock_rec = paginator.page(paginator.num_pages)

        create_fm = PurchaseCreationForm()
        update_fm = PurchaseUpdateForm()
        context = {'stock_rec': stock_rec, 'create_fm': create_fm, 'update_fm': update_fm}
        return render(request, 'admin__purchase_list.html', context)
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')
    

@admin_required
def create_purchase(request):
    try:
        form = PurchaseCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Record Added Successfully.....')
        else:
            error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
            messages.warning(request, error_message)
        return redirect('/admin/product_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

# def update_purchase(request): 
#     if request.method == 'POST':
#         purchase_id=request.POST.get('txt_id') 
#         purchase = get_object_or_404(Purchase, id=purchase_id)
#         form = PurchaseUpdateForm(request.POST, instance=purchase)
#         if form.is_valid():
#             form.save()
#             messages.success(request, 'Stock Updated Successfully.')
#             return redirect('/admin/purchase_list/')
#         else:
#             messages.warning(request, 'Form Not Submitted: Something Wrong')
#             return redirect('/admin/purchase_list/')
#     else:
#         messages.warning(request, 'Form Not Submited : Something Missing')
#         return redirect('/admin/purchase_list/')
    
@admin_required
def delete_purchase(request, id):
    try:
        purchase = get_object_or_404(Purchase, id=id)

        if purchase.quantity == 0:
            purchase.delete()
            messages.success(request, 'Stock Deleted Successfully.')
        else:
            qty = purchase.quantity
            stock = purchase.product.available_stock  # Assuming 'product' is a ForeignKey in Purchase model

            if qty > stock:
                messages.info(request, 'Quantity should be less than Stock')
                return redirect('/admin/purchase_list/')

            update_qty = int(stock) - int(qty)

            # Fetch the related Product instance and update its available_stock
            product_instance = purchase.product
            product_instance.available_stock = update_qty
            product_instance.save()

            purchase.delete()
            messages.success(request, 'Stock Deleted Successfully.')

        return redirect('/admin/purchase_list/')
    except Http404:
        return render(request, '404.html')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')
 
 
 

@admin_required
def invoice_list(request):
    try:
        
        rec = Invoice.objects.select_related().order_by('-id')
        Filter = Invoice_List_Filter(request.GET, queryset=rec)
        invoice_rec = Filter.qs 

        # Add pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(invoice_rec, 25)  # Show 10 invoices per page
        try:
            invoice_rec = paginator.page(page)
        except PageNotAnInteger:
            invoice_rec = paginator.page(1)
        except EmptyPage:
            invoice_rec = paginator.page(paginator.num_pages)

        create_invoice_fm = InvoiceCreationForm()
        update_invoice_fm = InvoiceUpdateForm()
        context = {
            'invoice_rec': invoice_rec,
            'create_invoice_fm': create_invoice_fm,
            'update_invoice_fm': update_invoice_fm,
            'filter': Filter
        }
        return render(request, 'admin__invoice_list.html', context)
    except Exception as e:
        return render(request, '404.html')
    

@admin_required
def create_invoice(request):
    try:
        form = InvoiceCreationForm(request.POST)
        invoice_type=request.POST.get('invoiceType') 
        if form.is_valid():
            saved_instance = form.save(commit=False)
            if invoice_type=='original':
                prefix = 'REDPL' 
                last_invoice = Invoice.objects.filter(is_original=True).select_related().order_by('-id').first()
                if last_invoice:
                    last_number = int(last_invoice.invoice_number[-6:])
                    next_invoice_number = last_number + 1
                else:
                    next_invoice_number = 1
                saved_instance.invoice_number = f'{prefix}{next_invoice_number:06d}'  # Associate the form with the specific invoice
                saved_instance.is_original = True  # Associate the form with the specific invoice
                saved_instance.save()
                saved_id = saved_instance.id 
                return redirect(f'/admin/invoice_item_list/{saved_id}')
            elif invoice_type=='dublicate':
                prefix = 'DBLDH'  
                last_invoice = Invoice.objects.filter(is_original=False).select_related().order_by('-id').first()
                if last_invoice:
                    last_number = int(last_invoice.invoice_number[-6:])
                    next_invoice_number = last_number + 1
                else: 
                    next_invoice_number = 1
                saved_instance.invoice_number = f'{prefix}{next_invoice_number:06d}'  # Associate the form with the specific invoice
                saved_instance.is_original = False  # Associate the form with the specific invoice
                saved_instance.save()
                saved_id = saved_instance.id 
                return redirect(f'/admin/invoice_item_list/{saved_id}')
        else:
            error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
            messages.warning(request, error_message)
        return redirect('/admin/invoice_list/')
    except Exception as e:
        return render(request, '404.html') 

@admin_required
def update_invoice(request): 
    try:
        if request.method == 'POST':
            invoice_id = request.POST.get('txt_id') 
            invoice = get_object_or_404(Invoice, id=invoice_id)
            form = InvoiceUpdateForm(request.POST, instance=invoice)
            if form.is_valid():
                form.save()
                messages.success(request, 'Invoice Updated Successfully.')
                return redirect('/admin/invoice_list/')
            else:
                error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
                messages.warning(request, error_message)
        else:
            messages.warning(request, 'Form Not Updated: Something Missing')
            return redirect('/admin/invoice_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')


@admin_required
def delete_invoice(request, id):
    try:
        invoice = get_object_or_404(Invoice, id=id)

        # Check if there are associated InvoiceItems
        if invoice.invoiceitem_set.exists():
            messages.warning(request, 'Invoice not deleted. Invoice is not empty.')
            return redirect('/admin/invoice_list/')

        # If no associated InvoiceItems, proceed with deletion
        with transaction.atomic():
            invoice.delete()
            messages.success(request, 'Invoice Deleted Successfully...')
            return redirect('/admin/invoice_list/')
    except Http404:
        return render(request, '404.html')

 
def add_extra_charges_in_invoice(request):
    id=request.session.get('session_invoice_id')
    if request.method == 'POST':
        packing_charges=request.POST.get('packing_charges')
        transportation_charges=request.POST.get('transportation_charges')
        Invoice.objects.filter(id=id).update(packing_charges=packing_charges,transportation_charges=transportation_charges,is_added_in_account=False)
        messages.success(request,'Record Updated Success...!')
    return redirect(f'/admin/invoice_item_list/{id}')



@admin_required
def add_invoice_in_account(request,id):
    try:
        invoice=get_object_or_404(Invoice, id=id)
        check_invoice_exist=Account.objects.filter(invoice_number=invoice.invoice_number)
        
        if not invoice.grand_total:
            messages.warning(request, 'Product not available.....') 
            return redirect('/admin/invoice_list/') 
        
        if invoice.packing_charges:
          packing_charges=invoice.packing_charges
        else:
          packing_charges=0

        if invoice.transportation_charges:
          transportation_charges=invoice.transportation_charges
        else:
          transportation_charges=0
 
        g_total_amount = InvoiceItem.objects.filter(invoice=invoice).aggregate(total_total_amount=Sum('total_amount'))['total_total_amount'] 
        if not g_total_amount:
            g_total_amount=0
        grand_total_amount=g_total_amount+packing_charges+transportation_charges
        
        Invoice.objects.filter(id=id).update(grand_total=grand_total_amount) 

        if check_invoice_exist.exists():
            check_invoice_exist.update(amount=grand_total_amount)
            Invoice.objects.filter(id=id).update(is_added_in_account=True)
            return redirect('/admin/invoice_list/')
        else:
            acc_rec_count = Account.objects.filter(dealer=invoice.dealer).count()
            if acc_rec_count:
                balance_rec = Account.objects.filter(dealer=invoice.dealer).latest('id')
                running_balance = balance_rec.balance - int(grand_total_amount) if acc_rec_count > 0 else -grand_total_amount
            else:
                running_balance=-int(grand_total_amount)

            Account.objects.create(
                dealer=invoice.dealer,
                amount=int(grand_total_amount),
                payment_mode='other',  # Add the appropriate payment mode
                is_credit=False,  # Adjust based on your logic
                balance=running_balance,
                invoice_number=invoice.invoice_number,
                created_by=request.user,  # Assuming you have a user associated with the request
            )
            Invoice.objects.filter(id=id).update(is_added_in_account=True)
            return redirect('/admin/invoice_list/') 
    except Exception as e:
            return render(request, '404.html')
        
@admin_required
def invoice_item_list(request, id):
    try:
        request.session['session_invoice_id'] = id
        records = InvoiceItem.objects.filter(invoice=id)
        sale_rec = Invoice.objects.get(id=id)
        form = InvoiceProductForm()

        if request.method == 'POST':
            form = InvoiceProductForm(request.POST)
 
            if form.is_valid():
                product_id = form.cleaned_data['product']
                product_qty = form.cleaned_data['quantity']
                available_stock = Product.objects.get(id=product_id.id).available_stock

                if int(available_stock) < int(product_qty):
                    messages.warning(request, f'Only {available_stock} Quantity Available')
                    return redirect(f'/admin/invoice_item_list/{id}')
                else:
                    fm = form.save(commit=False)
                    fm.invoice = sale_rec  # Associate the form with the specific invoice
                    fm.save()
                    messages.success(request, 'Item Added Successfully')
                    return redirect(f'/admin/invoice_item_list/{id}')
            else:
                # Form is not valid, handle the error
                messages.warning(request, 'Item Not Added. Please check the form.')

        total_quantity = InvoiceItem.objects.filter(invoice=sale_rec).aggregate(total_quantity=Sum('quantity'))['total_quantity']
        total_gst_amount = InvoiceItem.objects.filter(invoice=sale_rec).aggregate(total_gst_amount=Sum('gst_amount'))['total_gst_amount']
        sub_total_amount = InvoiceItem.objects.filter(invoice=sale_rec).aggregate(total_total_amount=Sum('total_amount'))['total_total_amount'] 
       
        if sale_rec.packing_charges:
          packing_charges=sale_rec.packing_charges
        else:
          packing_charges=0

        if sale_rec.transportation_charges:
          transportation_charges=sale_rec.transportation_charges
        else:
          transportation_charges=0
 
        g_total_amount = InvoiceItem.objects.filter(invoice=sale_rec).aggregate(total_total_amount=Sum('total_amount'))['total_total_amount'] 
        if not g_total_amount:
            g_total_amount=0
        grand_total_amount=g_total_amount+packing_charges+transportation_charges
        total_products = InvoiceItem.objects.filter(invoice=sale_rec).count()

        context = {'form': form, 'id': id, 'invoice_rec': records, 'total_quantity': total_quantity,
                   'total_gst_amount': total_gst_amount, 'grand_total_amount': grand_total_amount,'sub_total_amount':sub_total_amount,
                   'total_products': total_products, 'invoice_details': sale_rec}
        return render(request, 'admin__manage_invoice_item.html', context)

    except Invoice.DoesNotExist:
        # Handle the case when the invoice does not exist (404 Not Found)
        raise Http404("Invoice not found")

    except Product.DoesNotExist:
        # Handle the case when the product does not exist (404 Not Found)
        raise Http404("Product not found")

    except Exception as e:
        # Handle other exceptions
        messages.error(request, f'An error occurred: {str(e)}')
        return render(request, '404.html', status=404)

@admin_required
def delete_invoice_item(request, id):
    try:
        invoice = get_object_or_404(InvoiceItem, id=id)
        invoice.delete()
        id=request.session.get('session_invoice_id')
        messages.success(request, 'Product Deleted Successfully.')
        return redirect(f'/admin/invoice_item_list/{id}')
    except Http404:
        return render(request, '404.html')
 
@admin_required
def get_product_details(request):
    product_code = request.GET.get('product_code', '')
    product_id = request.GET.get('productId', '')  # Use 'product_id' to match the parameter in your AJAX request
    product = None 

    if product_code:
        product = Product.objects.filter(product_code=product_code).first()
    elif product_id:
        product = Product.objects.filter(id=product_id).first()
    if product:
        data = {
            'product_name': product.id,
            'sale_amount': product.sale_amount,
            'gst': product.gst,
            'available_stock': product.available_stock,
            # Include other product details as needed
        }
        return JsonResponse(data)
    else:
        return JsonResponse({'error': 'Product not found'})
    

import openpyxl

@admin_required
def dealer_bulk_creation(request):
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file')
        
        if excel_file:
            try:
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active
                data_to_insert = []
                num_records_inserted = 0

                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    dealer_name, business_name, mobile_no, email_id, address, state_name, pin_code, gst_number = row

                    # Check if the state exists
                    state_exist = State.objects.filter(state_name=state_name).exists()
                    if not state_exist:
                        messages.warning(request, f"{state_name} State Not Exist...!")
                        return redirect('/admin/dealer_list/')

                    # Fetch the existing state or create a new one if it doesn't exist
                    state_obj, created = State.objects.get_or_create(state_name=state_name)

                    # Create a dealer object
                    dealer_obj = Dealer(
                        dealer_name=dealer_name,
                        business_name=business_name,
                        mobile_no=mobile_no,
                        email_id=email_id,
                        address=address,
                        state=state_obj,
                        pin_code=pin_code,
                        gst_number=gst_number
                    )
                    data_to_insert.append(dealer_obj)
                    num_records_inserted += 1

                # Bulk create dealers
                Dealer.objects.bulk_create(data_to_insert)
                messages.success(request, f'{num_records_inserted} records inserted successfully')
                
            except openpyxl.utils.exceptions.CellCoordinatesException as e:
                messages.error(request, f'Error in Excel file format: {str(e)}')
            except Exception as e:
                messages.error(request, f'Error occurred during import: {str(e)}')
        else:
            messages.error(request, 'No file selected.')

    return redirect('/admin/dealer_list/')



@admin_required
def product_bulk_creation(request): 
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active 
                data_to_insert = []
                num_records_inserted = 0
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    product_code = row[0]
                    product_name = row[1]
                    description = row[2]
                    sale_amount = row[3]
                    hsn_sac = row[4]
                    gst = row[5]
                    available_stock = row[6]
                    minimum_stock = row[7]

                    # Check if product with the same code already exists
                    product_exist = Product.objects.filter(product_code=product_code).exists()
                    if product_exist:
                        messages.warning(request, f"Product with code {product_code} already exists.")
                        return redirect('/admin/product_list/')
                
                    num_records_inserted += 1

                    product_obj = Product(
                        product_code=product_code,
                        product_name=product_name,
                        description=description,
                        sale_amount=sale_amount,
                        hsn_sac=hsn_sac,
                        gst=gst,
                        available_stock=available_stock,
                        minimum_stock=minimum_stock
                    )
                    data_to_insert.append(product_obj)
                
                Product.objects.bulk_create(data_to_insert)
                messages.success(request, f'{num_records_inserted} records inserted successfully')
            except Exception as e:
                messages.error(request, f'Error occurred during import: {str(e)}')
        else:
            messages.error(request, 'No file selected.')
        return redirect('/admin/product_list/')

@admin_required
def print_all_invoice_formate(request, id):
    try:
        #This code for add add invoice amount in accont model
        invoice=get_object_or_404(Invoice, id=id)
        check_invoice_exist=Account.objects.filter(invoice_number=invoice.invoice_number)
        
        if not invoice.grand_total:
            messages.warning(request, 'Product not available.....') 
            return redirect('/admin/invoice_list/') 
        
        if invoice.packing_charges:
          packing_charges=invoice.packing_charges
        else:
          packing_charges=0

        if invoice.transportation_charges:
          transportation_charges=invoice.transportation_charges
        else:
          transportation_charges=0
 
        g_total_amount = InvoiceItem.objects.filter(invoice=invoice).aggregate(total_total_amount=Sum('total_amount'))['total_total_amount'] 
        if not g_total_amount:
            g_total_amount=0
        grand_total_amount=g_total_amount+packing_charges+transportation_charges
        Invoice.objects.filter(id=id).update(grand_total=grand_total_amount) 

        if check_invoice_exist.exists():
            check_invoice_exist.update(amount=grand_total_amount)
            Invoice.objects.filter(id=id).update(is_added_in_account=True)
        else:
            acc_rec_count = Account.objects.filter(dealer=invoice.dealer).count()
            if acc_rec_count:
                balance_rec = Account.objects.filter(dealer=invoice.dealer).latest('id')
                running_balance = balance_rec.balance - int(invoice.grand_total) if acc_rec_count > 0 else -int(invoice.grand_total)
            else:
                running_balance = -int(grand_total_amount)
            
            Account.objects.create(
                dealer=invoice.dealer,
                amount=int(invoice.grand_total),
                payment_mode='other',  # Add the appropriate payment mode
                is_credit=False,  # Adjust based on your logic
                balance=running_balance,
                invoice_number=invoice.invoice_number,
                created_by=request.user,  # Assuming you have a user associated with the request
            )
            Invoice.objects.filter(id=id).update(is_added_in_account=True)

        # Here code start for printing invoice
        invoice_details = Invoice.objects.get(id=id)
        item_rec = InvoiceItem.objects.filter(invoice=invoice_details)
        # Aggregate all values in a single query
        aggregates = item_rec.aggregate(
            total_quantity=Sum('quantity'),
            total_gst_amount=Sum('gst_amount'),
            grand_total_amount=Sum('total_amount'),
            total_rate=Sum('rate'),
            total_taxable_amount=Sum('taxable_amount'),
        )

        dealer_id = invoice_details.dealer.id 
        if invoice.packing_charges:
            packing_charges=invoice.packing_charges
        else:
            packing_charges=0

        if invoice.transportation_charges:
            transportation_charges=invoice.transportation_charges
        else:
            transportation_charges=0

            
        # Use the aggregates directly instead of querying the database again
        total_quantity = aggregates['total_quantity']
        total_gst_amount = aggregates['total_gst_amount']
        grand_total_amount = aggregates['grand_total_amount'] + int(packing_charges) + int(transportation_charges)
        total_rate = aggregates['total_rate']
        total_taxable_amount = aggregates['total_taxable_amount']
 
        latest_balance_record = Account.objects.filter(dealer=invoice.dealer).order_by('-id').first()

        if latest_balance_record:
            latest_balance = latest_balance_record.balance
        else:
            latest_balance=0

        gst_type = "IGST" if invoice_details.dealer.state != "Gujarat" else "CGST / SGST"
        context = {
            'invoice_details': invoice_details, 
            'item_rec': item_rec,
            'total_quantity': total_quantity,
            'total_gst_amount': total_gst_amount,
            'grand_total_amount': grand_total_amount,
            'total_rate': total_rate,
            'total_taxable_amount': total_taxable_amount,
            'gst_type': gst_type, 
            'latest_balance':latest_balance
        }
        return render(request, 'admin__print_all_invoice_formate.html', context)
    except Invoice.DoesNotExist:
        return render(request, '404.html')


@admin_required
def transaction_list(request):
    try: 
        rec = Account.objects.select_related().order_by('-id')
        last_account = rec.first()  # Correct usage of first(), without specifying 'id'
        if last_account:
            last_id = last_account.id 
        else:
            last_id=0
        Filter = Transaction_List_Filter(request.GET, queryset=rec)
        transaction_rec = Filter.qs 

        # Add pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(transaction_rec, 50)  # Show 10 transactions per page
        try:
            transaction_rec = paginator.page(page)
        except PageNotAnInteger:
            transaction_rec = paginator.page(1)
        except EmptyPage:
            transaction_rec = paginator.page(paginator.num_pages)
        latest_transaction = Account.objects.order_by('-id').first()
        latest_transaction_id = latest_transaction.id if latest_transaction else None
        create_transaction_fm = CreateTransactionForm()
        update_transaction_fm = UpdateTransactionForm() 
        context = {
            'latest_transaction_id': latest_transaction_id,
            'transaction_rec': transaction_rec,
            'create_transaction_fm': create_transaction_fm,
            'update_transaction_fm': update_transaction_fm,
            'filter': Filter,
            'last_id':last_id
        }
        return render(request, 'admin__transactions_list.html', context)
    except Exception as e:
        return render(request, '404.html')

@admin_required
def create_transaction(request):
    try:
        if request.method == 'POST':
            acc_rec_count = Account.objects.filter(dealer=request.POST.get('dealer')).count()
            if acc_rec_count:
                balance_rec = Account.objects.filter(dealer=request.POST.get('dealer')).latest('id')
            payment_status = request.POST.get('transaction-type-create')
            dealer_profile=Dealer.objects.get(id=int(request.POST.get('dealer')))
            form = CreateTransactionForm(request.POST)
            if form.is_valid():
                fm = form.save(commit=False)
                amount=int(request.POST.get('amount'))
                if payment_status == 'credit':
                    fm.is_credit=True
                    fm.balance = balance_rec.balance +  amount if acc_rec_count > 0 else amount
                else: 
                    fm.is_credit=False
                    fm.balance = balance_rec.balance -  amount if acc_rec_count > 0 else -amount
                fm.created_by=request.user
                fm.save()
                messages.success(request, 'Transaction Created Successfully.....')
            else:
                error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
                messages.warning(request, error_message)
            return redirect('/admin/transaction_list/')
        else:
            messages.warning(request, 'Something is missing ....!')
            return redirect('/admin/transaction_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

    
@admin_required
def update_transaction(request): 
    try:
        if request.method == 'POST':
            transaction_id = request.POST.get('txt_id') 
            update_payment_status = request.POST.get('transaction-type-update')
 
            transaction = get_object_or_404(Account, id=transaction_id)
            form = UpdateTransactionForm(request.POST, instance=transaction)
            if form.is_valid():
                fm = form.save(commit=False)
                if update_payment_status == 'credit':
                    fm.is_credit=True
                else:
                    fm.is_credit=False 

                fm.modified_by=request.user
                fm.save()
                messages.success(request, 'Trasaction Updated Successfully.')
                return redirect('/admin/troubleshoot_transactions_for_balance/')
            else:
                error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
                messages.warning(request, error_message)
                return redirect('/admin/transaction_list/')
        else:
            return redirect('/admin/transaction_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@admin_required
def delete_transaction(request, id):
    try:
        # trans_for_invoice = get_object_or_404(Account, id=id)
        # invoice = get_object_or_404(Invoice, invoice_number=trans_for_invoice.invoice_number)
        # Invoice.objects.filter(id=invoice.id).update(is_added_in_account=False)

        transaction = get_object_or_404(Account, id=id)
        transaction.delete()
        messages.success(request, 'Transaction Deleted Successfully.')
        return redirect('/admin/troubleshoot_transactions_for_balance/')
    except Http404:
        return render(request, '404.html')
 
@admin_required
def print_transactions_receipt(request,id):
    try: 
        data=Account.objects.get(id=id)
        context={'data':data}
        return render(request, 'admin__print_payment_received_slip.html', context)
    except Exception as e:
        return render(request, '404.html')

@admin_required
def troubleshoot_transactions_for_balance(request):
    try:
        dealers = Dealer.objects.select_related()

        for dealer in dealers:
            # Filter previous entries for the same dealer
            previous_entries = Account.objects.filter(dealer=dealer).order_by('transaction_date')

            for account in previous_entries:
                # Calculate previous credited and debited amounts
                previous_credited_amount = previous_entries.filter(
                    transaction_date__lte=account.transaction_date, is_credit=True
                ).aggregate(Sum('amount'))['amount__sum'] or 0

                previous_debited_amount = previous_entries.filter(
                    transaction_date__lte=account.transaction_date, is_credit=False
                ).aggregate(Sum('amount'))['amount__sum'] or 0

                # Update the balance field 
                account.balance = previous_credited_amount - previous_debited_amount
                account.save() 
                    

        return redirect('/admin/transaction_list/')
    except Exception as e:
        return render(request, '404.html')
    


@admin_required
def performa_invoice_list(request):
    try:
        rec = PerformaInvoice.objects.select_related().order_by('-id')
        Filter=Performa_Invoice_List_Filter(request.GET, queryset=rec)
        invoice_rec=Filter.qs 
        create_invoice_fm = PerformaInvoiceCreationForm()
        update_invoice_fm = PerformaInvoiceUpdateForm()
        context = {'invoice_rec': invoice_rec, 'create_invoice_fm': create_invoice_fm, 'update_invoice_fm': update_invoice_fm,'filter':Filter}
        return render(request, 'admin__performa_invoice_list.html', context)
    except Exception as e:
        return render(request, '404.html')

@admin_required
def create_performa_invoice(request):
    try:
        form = PerformaInvoiceCreationForm(request.POST)
        if form.is_valid():
            saved_instance = form.save()
            saved_id = saved_instance.id
            return redirect(f'/admin/performa_invoice_item_list/{saved_id}')
        else:
            error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
            messages.warning(request, error_message)
        return redirect('/admin/performa_invoice_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@admin_required
def update_performa_invoice(request): 
    try:
        if request.method == 'POST':
            performa_invoice_id = request.POST.get('txt_id') 
            performa_invoice = get_object_or_404(PerformaInvoice, id=performa_invoice_id)
            form = PerformaInvoiceUpdateForm(request.POST, instance=performa_invoice)
            if form.is_valid():
                form.save()
                messages.success(request, 'Invoice Updated Successfully.')
                return redirect('/admin/performa_invoice_list/')
            else:
                error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
                messages.warning(request, error_message)
        else:
            messages.warning(request, 'Form Not Updated: Something Missing')
            return redirect('/admin/invoice_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')


@admin_required
def delete_performa_invoice(request, id):
    try:
        performa_invoice = get_object_or_404(PerformaInvoice, id=id)
        with transaction.atomic():
            performa_invoice.delete()
            messages.success(request, 'Performa Invoice Deleted Successfully...')
            return redirect('/admin/performa_invoice_list/')
    except Http404:
        return render(request, '404.html')

     
@admin_required
def performa_invoice_item_list(request, id):
    try:
        request.session['session_invoice_id'] = id
        records = PerformaInvoiceItem.objects.filter(performa_invoice=id)
        sale_rec = PerformaInvoice.objects.get(id=id)
        form = PerformaInvoiceProductForm()

        if request.method == 'POST':
            form = PerformaInvoiceProductForm(request.POST)

            if form.is_valid():
                fm = form.save(commit=False)
                fm.performa_invoice = sale_rec
                fm.save()
                messages.success(request, 'Item Added Successfully')
                return redirect(f'/admin/performa_invoice_item_list/{id}')
            else:
                # Form is not valid, handle the error
                messages.warning(request, 'Item Not Added. Please check the form.')

        total_quantity = PerformaInvoiceItem.objects.filter(performa_invoice=sale_rec).aggregate(total_quantity=Sum('quantity'))['total_quantity']
        total_gst_amount = PerformaInvoiceItem.objects.filter(performa_invoice=sale_rec).aggregate(total_gst_amount=Sum('gst_amount'))['total_gst_amount']
        grand_total_amount = PerformaInvoiceItem.objects.filter(performa_invoice=sale_rec).aggregate(total_total_amount=Sum('total_amount'))['total_total_amount']
        total_products = PerformaInvoiceItem.objects.filter(performa_invoice=sale_rec).count()

        context = {'form': form, 'id': id, 'invoice_rec': records, 'total_quantity': total_quantity,
                   'total_gst_amount': total_gst_amount, 'grand_total_amount': grand_total_amount,
                   'total_products': total_products, 'invoice_details': sale_rec}
        return render(request, 'admin__performa_invoice_item.html', context)

    except PerformaInvoice.DoesNotExist:
        # Handle the case when the Performa invoice does not exist (404 Not Found)
        messages.error(request, 'Performa Invoice not found.')
        return redirect('/admin/performa_invoice_list/')

    except Exception as e:
        # Handle other exceptions
        messages.error(request, f'An error occurred: {str(e)}')
        return render(request, '404.html', status=404)

@admin_required
def delete_performa_invoice_item(request, id):
    try:
        invoice = get_object_or_404(PerformaInvoiceItem, id=id)
        invoice.delete()
        id=request.session.get('session_invoice_id')
        messages.success(request, 'Product Deleted Successfully.')
        return redirect(f'/admin/performa_invoice_item_list/{id}')
    except Http404:
        return render(request, '404.html')


@admin_required
def print_performa_invoice(request, id):
    try:
        invoice_details = PerformaInvoice.objects.get(id=id)
        item_rec = PerformaInvoiceItem.objects.filter(performa_invoice=id)
         
        total_quantity = PerformaInvoiceItem.objects.filter(performa_invoice=invoice_details).aggregate(total_quantity=Sum('quantity'))['total_quantity']
        total_gst_amount = PerformaInvoiceItem.objects.filter(performa_invoice=invoice_details).aggregate(total_gst_amount=Sum('gst_amount'))['total_gst_amount']
        grand_total_amount = PerformaInvoiceItem.objects.filter(performa_invoice=invoice_details).aggregate(total_total_amount=Sum('total_amount'))['total_total_amount']
        total_rate = item_rec.aggregate(total_rate=Sum('rate'))['total_rate']
        total_taxable_amount = item_rec.aggregate(total_taxable_amount=Sum('taxable_amount'))['total_taxable_amount']

        if not str(invoice_details.state) == "Gujarat":
            gst_type="IGST"
        else: 
            gst_type="CGST / SGST"
 

        context = {'invoice_details': invoice_details, 
                   'item_rec': item_rec,
                   'total_quantity':total_quantity,
                   'total_gst_amount':total_gst_amount,
                   'grand_total_amount':grand_total_amount,
                   'total_rate':total_rate,
                   'total_taxable_amount':total_taxable_amount,
                   'gst_type':gst_type,
        }
        return render(request, 'admin__print_performa_invoice.html', context)
    except Invoice.DoesNotExist:
        return render(request, '404.html')
    